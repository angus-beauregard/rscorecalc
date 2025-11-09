# dashboard.py
import datetime
import os
import pandas as pd
import streamlit as st

# ---- try to import your OCR stub; if it's not there, make a dummy one ----
try:
    from ocr_engine import extract_courses_from_image
except Exception:
    def extract_courses_from_image(_bytes):
        return pd.DataFrame(
            columns=["course_name", "mark", "group_avg", "group_sd", "credits"]
        )

EXCEL_RSCORE_FILE = "R-Score Calculator (perfect).xlsx"
IDGZ_CSV_PATH = "idgz+isgz_data.csv"   # high school / board data file


# ---------------------------------------------------------
# load high school / board table
# ---------------------------------------------------------
def load_idgz_table(path: str = IDGZ_CSV_PATH) -> pd.DataFrame:
    """Normalize the uploaded idgz+isgz_data.csv so we always return: school, isgz, idgz"""
    if not os.path.exists(path):
        return pd.DataFrame({"school": ["(default)"], "isgz": [0.0], "idgz": [1.0]})

    df = pd.read_csv(path)
    lower_map = {c: c.strip().lower() for c in df.columns}

    def find_col(candidates):
        for cand in candidates:
            if cand in lower_map.values():
                return [orig for orig, low in lower_map.items() if low == cand][0]
        return None

    school_col = find_col(["school board", "school", "high school", "board", "school name"])
    isgz_col = find_col(["isgz estimate", "isgz", "isg", "isgz_estimate"])
    idgz_col = find_col(["idgz estimate", "idgz", "idg", "idgz_estimate"])

    norm = pd.DataFrame()
    norm["school"] = df[school_col] if school_col else ["(default)"]
    norm["isgz"] = df[isgz_col] if isgz_col else 0.0
    norm["idgz"] = df[idgz_col] if idgz_col else 1.0
    return norm


# ---------------------------------------------------------
# optional excel params (kept but not required)
# ---------------------------------------------------------
def load_excel_rscore_params(path: str):
    base_constant = 35
    isg_avg = 0.0
    if not os.path.exists(path):
        return {"base": base_constant, "isg_avg": isg_avg}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        if "ISG Range" in wb.sheetnames:
            ws = wb["ISG Range"]
            isg_avg = ws["B10"].value or 0.0
        return {"base": base_constant, "isg_avg": float(isg_avg)}
    except Exception:
        return {"base": base_constant, "isg_avg": 0.0}


RSCORE_PARAMS = load_excel_rscore_params(EXCEL_RSCORE_FILE)
BASE_CONST = RSCORE_PARAMS["base"]
ISG_AVG = RSCORE_PARAMS["isg_avg"]


# ---------------------------------------------------------
# R-score math
# ---------------------------------------------------------
def compute_rscore_school_based(
    mark: float,
    group_avg: float,
    group_sd: float,
    idgz: float = 1.0,
    isgz: float = 0.0,
    C: float = 35.0,
    D: float = 1.0,
) -> float:
    if group_sd is None or group_sd == 0:
        z = 0.0
    else:
        z = (mark - group_avg) / group_sd
    r = ((z * idgz) + isgz + C) * D
    return round(r, 2)


def compute_overall_rscore(df: pd.DataFrame) -> float:
    if df.empty:
        return 0.0
    df = df.copy()
    if "credits" not in df.columns:
        df["credits"] = 1.0
    total_credits = df["credits"].sum()
    if total_credits == 0:
        return 0.0
    weighted = (df["rscore"] * df["credits"]).sum() / total_credits
    return round(weighted, 2)
    
from supabase import create_client
supabase = create_client(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_KEY"))

def save_rscore_to_db(user_email: str, rscore: float):
    """Save or update the user's current R-score."""
    if not user_email:
        return
    try:
        supabase.table("user_rscores").upsert(
            {"email": user_email, "rscore": rscore, "timestamp": datetime.datetime.now().isoformat()}
        ).execute()
    except Exception as e:
        st.warning(f"Could not save R-score: {e}")

def rank_courses_to_improve(
    df: pd.DataFrame,
    bump: int = 5,
    idgz: float = 1.0,
    isgz: float = 0.0,
) -> pd.DataFrame:
    base = compute_overall_rscore(df)
    rows = []
    for idx, row in df.iterrows():
        tmp = df.copy()
        improved_mark = min(row["mark"] + bump, 100)
        tmp.loc[idx, "rscore"] = compute_rscore_school_based(
            improved_mark,
            row["group_avg"],
            row["group_sd"],
            idgz=idgz,
            isgz=isgz,
        )
        new_overall = compute_overall_rscore(tmp)
        diff = round(new_overall - base, 3)
        rows.append(
            {
                "course_name": row["course_name"],
                "current_mark": row["mark"],
                f"overall_gain_if_+{bump}": diff,
            }
        )
    return pd.DataFrame(rows).sort_values(f"overall_gain_if_+{bump}", ascending=False)


# ---------------------------------------------------------
# forgiving CSV parser
# ---------------------------------------------------------
def normalize_col_name(raw: str) -> str:
    s = raw.strip().lower()
    s = s.replace(".", " ").replace("_", " ").replace("-", " ")
    s = " ".join(s.split())

    if any(k in s for k in ["course", "class name", "subject"]) or s == "class":
        return "course_name"
    if "grade" in s or "mark" in s or "note" in s or "result" in s or "score" in s:
        return "mark"
    if "avg" in s or "average" in s or "class avg" in s or "group avg" in s:
        return "group_avg"
    if "std" in s or "sd" in s or "standard" in s or "écart" in s or "ecart" in s:
        return "group_sd"
    if "credit" in s or "unit" in s or "unité" in s or "unites" in s:
        return "credits"
    return s


def coerce_csv_columns(df: pd.DataFrame) -> pd.DataFrame:
    col_map = {col: normalize_col_name(col) for col in df.columns}
    df = df.rename(columns=col_map)

    needed = ["course_name", "mark", "group_avg", "group_sd", "credits"]
    missing = [c for c in needed if c not in df.columns]

    if "credits" in missing:
        df["credits"] = 1
        missing.remove("credits")

    if missing:
        raise ValueError(
            f"CSV is missing these columns even after normalization: {missing}. "
            "We accept headers like: Course Name, Your Grade, Class Avg, Std. Dev, Credits."
        )

    return df[needed]


# ---------------------------------------------------------
# main Streamlit page
# ---------------------------------------------------------
def show_dashboard():
    # session defaults
    if "courses" not in st.session_state:
        st.session_state["courses"] = pd.DataFrame(
            columns=["course_name", "mark", "group_avg", "group_sd", "credits"]
        )
    if "target_r" not in st.session_state:
        st.session_state["target_r"] = 30.0
    if "semester_end" not in st.session_state:
        st.session_state["semester_end"] = datetime.date(2025, 12, 19)
    if "semester_start" not in st.session_state:
        st.session_state["semester_start"] = datetime.date(2025, 8, 25)

    hs_df = load_idgz_table()

    st.markdown("## 📈 RScore Premium Dashboard")
    st.markdown("Upload your grades, we’ll apply your school factors, and show your R-score.")

    left, right = st.columns([1.2, 0.8])

    # ---------------- LEFT: TABS ----------------
    with left:
        tabs = st.tabs(["📷 OCR upload", "📄 CSV upload", "✍️ Manual entry", "📈 Biggest gains", "🎯 Goals"])

        # --- OCR upload ---
        with tabs[0]:
            st.markdown("#### Upload screenshot (OCR)")
            img_file = st.file_uploader("Drop an image here", type=["png", "jpg", "jpeg"], key="ocr_upload")
            if img_file is not None:
                img_bytes = img_file.read()
                df_ocr = extract_courses_from_image(img_bytes)
                st.success("OCR processed.")
                st.dataframe(df_ocr, hide_index=True)
                st.session_state["courses"] = df_ocr

        # --- CSV upload ---
        with tabs[1]:
            st.markdown("#### Upload CSV")
            st.caption("We accept: Course Name, Your Grade, Class Avg, Std. Dev, Credits (any order, any case).")
            csv_file = st.file_uploader("Upload CSV", type=["csv"], key="csv_upload")
            if csv_file is not None:
                try:
                    df_csv = pd.read_csv(csv_file)
                    df_csv = coerce_csv_columns(df_csv)
                except Exception as e:
                    st.error(f"Could not understand your CSV: {e}")
                else:
                    st.session_state["courses"] = df_csv

            editable_df = st.data_editor(
                st.session_state["courses"], num_rows="dynamic", key="csv_editor"
            )
            st.session_state["courses"] = editable_df

            if not editable_df.empty:
                st.download_button(
                    "Download results (CSV)",
                    data=editable_df.to_csv(index=False),
                    file_name="rscore_results.csv",
                    mime="text/csv",
                )

        # --- Manual entry ---
        with tabs[2]:
            st.markdown("#### Add a course manually")
            with st.form("manual_form"):
                cname = st.text_input("Course name")
                mark = st.number_input("Your mark (%)", 0.0, 100.0, 85.0)
                gavg = st.number_input("Class / group average (%)", 0.0, 100.0, 75.0)
                gsd = st.number_input("Std. dev.", 0.0, 30.0, 8.0)
                creds = st.number_input("Credits", 0.5, 8.0, 2.0)
                submitted = st.form_submit_button("Add course")
            if submitted:
                new_row = {
                    "course_name": cname,
                    "mark": mark,
                    "group_avg": gavg,
                    "group_sd": gsd,
                    "credits": creds,
                }
                st.session_state["courses"] = pd.concat(
                    [st.session_state["courses"], pd.DataFrame([new_row])],
                    ignore_index=True,
                )
                st.success(f"Added {cname}")

        # --- Biggest gains ---
        with tabs[3]:
            st.markdown("#### Biggest gains")
            st.write("We simulate improving each course and show which one increases your overall R-score the most.")
            bump_amount = st.number_input(
                "Simulate improving each course by this many points:",
                min_value=1, max_value=15, value=5, step=1, key="bump_in_gains"
            )

            df_current = st.session_state["courses"].copy()

            if df_current.empty:
                st.info("Add or upload courses first.")
            else:
                # get selected high school + factors
                selected_hs = st.session_state.get("selected_hs", hs_df.iloc[0]["school"])
                hs_row = hs_df[hs_df["school"] == selected_hs].iloc[0]
                idgz_val = float(hs_row.get("idgz", 1.0))
                isgz_val = float(hs_row.get("isgz", 0.0))

                df_current["rscore"] = df_current.apply(
                    lambda row: compute_rscore_school_based(
                        row["mark"], row["group_avg"], row["group_sd"],
                        idgz=idgz_val, isgz=isgz_val
                    ),
                    axis=1,
                )

                gain_df = rank_courses_to_improve(df_current, bump=bump_amount,
                                                  idgz=idgz_val, isgz=isgz_val)

                st.success(f"R-score gains simulated using **{selected_hs}** (IDGZ={idgz_val}, ISGZ={isgz_val}).")
                st.caption("Higher values indicate the biggest potential boost to your semester R-score.")
                st.dataframe(gain_df, hide_index=True, use_container_width=True)

        # --- Goals ---
        with tabs[4]:
            st.markdown("#### Goals")
            st.write("Set your target R-score. The overview on the right uses this.")
            new_target = st.number_input(
                "Target R-score",
                min_value=0.0, max_value=50.0,
                value=st.session_state["target_r"], step=0.1,
                key="target_input_goals",
            )
            st.session_state["target_r"] = new_target

       # ---------------- RIGHT PANEL ----------------
    with right:
        school_options = hs_df["school"].tolist()
        selected_hs = st.selectbox(
            "High school / board (for IDGZ & ISGZ)",
            options=school_options,
            index=(
                school_options.index(st.session_state.get("selected_hs", school_options[0]))
                if st.session_state.get("selected_hs", None) in school_options
                else 0
            ),
            key="school_selector"
        )

        # 👇 Force a refresh when the school changes
        if st.session_state.get("_last_school") != selected_hs:
            st.session_state["_last_school"] = selected_hs
            st.rerun()

        st.session_state["selected_hs"] = selected_hs

        hs_row = hs_df[hs_df["school"] == selected_hs].iloc[0]
        school_idgz = float(hs_row.get("idgz", 1.0))
        school_isgz = float(hs_row.get("isgz", 0.0))

        df = st.session_state["courses"].copy()
        if not df.empty:
            df["rscore"] = df.apply(
                lambda row: compute_rscore_school_based(
                    row["mark"],
                    row["group_avg"],
                    row["group_sd"],
                    idgz=school_idgz,
                    isgz=school_isgz
                ),
                axis=1,
            )
            overall = compute_overall_rscore(df)
        else:
            overall = 0.0

        gap = round(st.session_state["target_r"] - overall, 2)

        col_over, col_sem = st.columns(2)

        with col_over:
            st.markdown("### Overview")
            st.markdown("Current R-score")
            st.markdown(f"<h2 style='margin-top:0'>{overall}</h2>", unsafe_allow_html=True)
            st.markdown("Gap to target")
            if df.empty:
                st.markdown("<h2 style='margin-top:0'>—</h2>", unsafe_allow_html=True)
            elif gap <= 0:
                st.markdown("<h2 style='margin-top:0'>On target ✅</h2>", unsafe_allow_html=True)
            else:
                st.markdown(f"<h2 style='margin-top:0'>+{gap}</h2>", unsafe_allow_html=True)

                with col_sem:
                    st.markdown("### Semester countdown")
                    semester_end_dates = {
                        "John Abbott College": datetime.date(2025, 12, 19),
                        "Marianopolis College": datetime.date(2025, 12, 18),
                        "Dawson College": datetime.date(2025, 12, 19),
                        "Vanier College": datetime.date(2025, 12, 19),
                        "Champlain (St-Lambert)": datetime.date(2025, 12, 19),
                        "Other / custom": None,
                    }
                    cegep = st.selectbox("Select cégep", list(semester_end_dates.keys()))
                    end_date = semester_end_dates[cegep] or st.date_input(
                        "Pick semester end", value=st.session_state["semester_end"]
                    )
                    st.session_state["semester_end"] = end_date

                    today = datetime.date.today()
                    start_date = st.session_state["semester_start"]
                    total_days = (end_date - start_date).days
                    days_left = (end_date - today).days
                    days_done = (today - start_date).days

                    if total_days > 0:
                        percent = min(max(int((days_done / total_days) * 100), 0), 100)
                    else:
                        percent = 0

                    st.progress(percent / 100)
                    if days_left >= 0:
                        st.write(f"📅 **{days_left} days** left in the semester ({percent}%)")
                    else:
                        st.write("📅 Semester has ended")





